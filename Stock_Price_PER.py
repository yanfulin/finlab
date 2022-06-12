import logging

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from datetime import date
import sqlite3

# 定義輸出格式
FORMAT = '%(asctime)s %(filename)s %(levelname)s:  %(message)s'
# Logging初始設定 + 上定義輸出格式
logging.basicConfig(level=logging.DEBUG, format=FORMAT)

def GetDividend(stock_id):

    conn = sqlite3.connect('fin_db')  # 建立連線
    sql_query = f'select stock_id, max(權利分派基準日),"盈餘分配之股東現金股利(元/股)", "法定盈餘公積、資本公積發放之現金(元/股)"  FROM "dividend_announcement" where stock_id={stock_id};'
    records = pd.read_sql(sql_query,conn)  # 取得回傳資料
    conn.close()
    logging.debug(f'{records}')
    dividend=records["盈餘分配之股東現金股利(元/股)"][0]+records["法定盈餘公積、資本公積發放之現金(元/股)"][0]
    logging.debug(f'{stock_id} dividiend of the lastest record is {dividend}')
    return dividend


def GetPER(stock_id):

    conn = sqlite3.connect('fin_db')  # 建立連線
    records = pd.read_sql('SELECT * FROM "price_earning_ratio-本益比";',conn)  # 取得回傳資料
    conn.close()


    # Select PER for the past 3 year
    # 02d is to make 2020-1-1 format 2020-01-01
    today = pd.to_datetime('today')
    start_date=f"{today.year-3}-{today.month:02d}-{today.day:02d}"
    # print(start_date)
    condition = (records['date'] > start_date)
    records=records.loc[condition]
    # print(records)
    # print(records[stock_id])
    # print("min PER is", records[stock_id].min())
    # print("avg PER is ",records[stock_id].mean())
    max_PER=float(records[stock_id].max())
    min_PER=float(records[stock_id].min())
    avg_PER=float(records[stock_id].mean())
    return max_PER, min_PER, avg_PER

def GetEPS(stock_id):
    max_PER, min_PER, avg_PER = GetPER(stock_id)
    print(f"Get the EPS of {stock_id} max {max_PER}, min {min_PER}, average {avg_PER}")

    conn = sqlite3.connect('fin_db')  # 建立連線
    sql_query=f'SELECT date,"{stock_id}" FROM "EPS_forecast_table";'
    logging.debug(f'GETEPS "{stock_id}" via SQL      {sql_query}')
    df_EPS = pd.read_sql(sql_query, conn)  # 取得回傳資料
    conn.close()

    today = pd.to_datetime('today')

    df_EPS["year"]=df_EPS["date"].map(lambda x: x.split('-')[0]).astype(int)

    EPS_yearly = df_EPS.groupby("year")[stock_id].sum()

    EPS_yearly = pd.DataFrame({'Year': EPS_yearly.index, 'EPS': EPS_yearly.values})


    EPS_yearly["max_stock_price"]= (max_PER * EPS_yearly["EPS"])
    EPS_yearly["avg_stock_price"] = (avg_PER * EPS_yearly["EPS"])
    EPS_yearly["min_stock_price"] = (min_PER * EPS_yearly["EPS"])
    #print(EPS_yearly)
    logging.debug(f'{stock_id} Yearly EPS {EPS_yearly}')

    #EPS_yearly.to_csv(EPS_yearly_file)
    previous_year=today.year-1
    this_year=today.year
    #print("three year EPS=====",EPS_yearly.loc[(EPS_yearly["Year"]<=(this_year)),"EPS"])
    three_year_EPS=EPS_yearly.loc[(EPS_yearly["Year"]<=(this_year)),"EPS"].values[-3:]


    logging.debug(f'{stock_id} three year EPS {three_year_EPS}')
    return three_year_EPS




def GetEPS_per_quauter(stock_id):
    conn = sqlite3.connect('fin_db')  # 建立連線
    df_EPS = pd.read_sql('SELECT * FROM "financial_statement-每股盈餘";', conn)  # 取得回傳資料
    print(df_EPS.head())


    df_EPS["year"] = df_EPS["date"].map(lambda x: x.split('-')[0]).astype(int)
    today = pd.to_datetime('today')
    previous_year=today.year-2
    two_year_EPS=df_EPS.loc[(df_EPS["year"]>=previous_year)][stock_id][::-1].values
    size = len(two_year_EPS)
    # print(two_year_EPS)
    # for n in range(size):
    #     print (f'size is {size}, n is {n}, 24+n is {24+n}, size-n-1 is {size-n-1}')
    #     print(24+n, size-n-1, two_year_EPS[size-n-1])

    #print(two_year_EPS)
    logging.debug(f'{stock_id} two year EPS per quarter {two_year_EPS}')
    return two_year_EPS, size


def export_to_excel(stock_id):
    Excel_file = Path.cwd() / "stocks.xlsx"
    wb = load_workbook(Excel_file)
    sheet = wb['evaluation']
    # data = sheet.values
    # data =list(data)
    # stock_list2 = [r[0] for r in data]
    # print(stock_list2)



def export_to_excel2():

    Excel_file = Path.cwd() / "stocks.xlsx"
    wb = load_workbook(Excel_file)
    sheet = wb['evaluation']
    #print (sheet.values)
    colnames = ["Stock ID","Chinese Name","Name","Current Price","52W High","52W Low","Beta","Market Cap","PBR","PER",
                "Target Price","Rick's PER","Comment","短中長投資屬性",
                "Dividend","Yield %","Previous Year EPS","This Year EPS","Next Year EPS", "Max PER","Min PER","Avg PER",
                "Price @ High PER","Price @ Avg PER","Price @ Min PER",
                "2010 Q1 EPS", "2010 Q2 EPS", "2010 Q3 EPS", "2010 Q4 EPS",
                "2011 Q1 EPS","2011 Q2 EPS","2011 Q3 EPS","2011 Q4 EPS",
                "2012 Q1 EPS","2012 Q2 EPS","2012 Q3 EPS","2012 Q4 EPS",
                "2010 Total EPS","2011 Total EPS","2012 Total EPS",
                "過去五年平均配息％","股利", "Yield %", "PER High","PER Low","Price @High PE", "Price @Low PE",
                "Price@Low PE- Market Price", "Price @6% Yield"]
    col_indices = {n: cell.value for n, cell in enumerate(sheet['1']) if cell.value in colnames}
    print (col_indices)
    stock_list=[]
    for row in sheet["A"]:
            if row.value != None:
                #print ("row===>", row.value)
                stock_list.append(row.value)
    print(stock_list)
    # stock list: some are in text. some are int. Needs to convert them to right format
    # Get the corresponding EPS and PER and write them into the right position
    for index, stock in enumerate(stock_list, start=1):
        #print(index, stock)
        if index ==1:
            pass
        else:
            stock = str(stock)
            #print(stock)
            EPS = GetEPS(stock)
            PER_max, PER_avg, PER_min = GetPER(stock)
            EPS_quarterly, size=GetEPS_per_quauter(stock)
            sheet.cell(row=index, column=16).value = GetDividend(stock)
            sheet.cell(row=index, column=18).value = EPS_quarterly[-4:].sum()
            sheet.cell(row=index, column=19).value = EPS[1]
            sheet.cell(row=index, column=20).value = EPS[2]
            sheet.cell(row=index, column=21).value = PER_max
            sheet.cell(row=index, column=22).value = PER_avg
            sheet.cell(row=index, column=23).value = PER_min

            #Write quarterly EPS. From 2020 Q1 to now (2011)
            for n in range(size):
                sheet.cell(row=index, column=(27+n)).value = EPS_quarterly[size-n-1]


            ### Save dataframe into excel file.

            #file_name2 = f'stocks_{date.today()}.xlsx'
            #wb.save(Path.cwd() / "stock_excel_output" / file_name2)
            file_name = f'stocks.xlsx'
            wb.save(Path.cwd()/file_name)







def main():
    fin = open('StockCode', 'r+')
    StockCodeList = [str(i) for i in fin.read().splitlines()]
    fin.close()
    # print(StockCodeList)

    # for ID in StockCodeList:
    #     print("ID=", ID)
    #     max_PER, min_PER, avg_PER = GetPER(ID)
    #     EPS = GetEPS(ID)
    #     #print (ID, max_PER, min_PER, avg_PER)
    #     print ("2020 EPS=" , EPS)
    export_to_excel2()



if __name__ == "__main__":
    main()
    # Test function by function
    # GetPER("2330")
    # GetEPS("2330")
    # GetEPS_per_quauter("2330")
    #GetDividend("2382")